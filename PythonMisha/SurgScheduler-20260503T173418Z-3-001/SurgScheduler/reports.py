import os
import pandas as pd
import matplotlib

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from datetime import datetime
import database as db

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

PALETTE = ["#1a6b8a", "#2eafd4", "#e8a838", "#e05252", "#4caf82", "#9b59b6"]
sns.set_theme(style="darkgrid", palette=PALETTE)

def export_surgeries_csv() -> str:
    rows = db.get_all_surgeries_df_data()
    df = pd.DataFrame(rows)
    cols = ["id","patient_name","surgeon_name","surgery_type","scheduled_date",
            "scheduled_time","duration_min","status","priority","room_name","notes"]
    df = df[[c for c in cols if c in df.columns]]
    df.columns = [c.replace("_"," ").title() for c in df.columns]
    path = os.path.join(REPORTS_DIR, f"surgeries_{_stamp()}.csv")
    df.to_csv(path, index=False)
    return path


def export_surgeries_excel() -> str:
    rows = db.get_all_surgeries_df_data()
    df = pd.DataFrame(rows)
    cols = ["id","patient_name","surgeon_name","surgery_type","scheduled_date",
            "scheduled_time","duration_min","status","priority","room_name","notes"]
    df = df[[c for c in cols if c in df.columns]]
    df.columns = [c.replace("_"," ").title() for c in df.columns]

    path = os.path.join(REPORTS_DIR, f"surgeries_{_stamp()}.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Surgeries", index=False)
        ws = writer.sheets["Surgeries"]
        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        # Header style
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1a6b8a")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
    return path


def export_patients_excel() -> str:
    rows = db.get_all_patients()
    df = pd.DataFrame(rows)
    df.columns = [c.replace("_"," ").title() for c in df.columns]
    path = os.path.join(REPORTS_DIR, f"patients_{_stamp()}.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Patients", index=False)
        ws = writer.sheets["Patients"]
        from openpyxl.styles import PatternFill, Font, Alignment
        hf = PatternFill("solid", fgColor="2eafd4")
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ml = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)
    return path


def export_surgeons_excel() -> str:
    rows = db.get_all_surgeons()
    df = pd.DataFrame(rows)
    df.columns = [c.replace("_"," ").title() for c in df.columns]
    path = os.path.join(REPORTS_DIR, f"surgeons_{_stamp()}.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Surgeons", index=False)
    return path


# ════════════════════════════════════════════════════════════════════════════
#  GRAPHICAL REPORTING  (Matplotlib / Seaborn)
# ════════════════════════════════════════════════════════════════════════════

def _base_fig(title: str, figsize=(9, 5)):
    fig, ax = plt.subplots(figsize=figsize, facecolor="#0f1923")
    ax.set_facecolor("#162433")
    ax.tick_params(colors="#a0b8c8")
    ax.xaxis.label.set_color("#a0b8c8")
    ax.yaxis.label.set_color("#a0b8c8")
    for spine in ax.spines.values():
        spine.set_edgecolor("#2a3f52")
    fig.suptitle(title, color="#e8f4f8", fontsize=14, fontweight="bold", y=0.98)
    return fig, ax


def chart_surgeries_by_status(parent_frame=None):
    stats = db.get_surgeries_stats()
    data = stats["by_status"]
    if not data:
        _show_no_data(); return None
    labels, values = zip(*data.items())
    colors = ["#4caf82","#e8a838","#2eafd4","#e05252"]
    fig, ax = _base_fig("Surgeries by Status")
    wedges, texts, autotexts = ax.pie(
        values, labels=labels, colors=colors[:len(labels)],
        autopct="%1.1f%%", startangle=140,
        textprops={"color": "#e8f4f8", "fontsize": 10},
        wedgeprops={"edgecolor": "#0f1923", "linewidth": 2}
    )
    for at in autotexts:
        at.set_fontsize(9)
    ax.set_aspect("equal")
    plt.tight_layout()
    return fig


def chart_surgeries_by_surgeon(parent_frame=None):
    stats = db.get_surgeries_stats()
    rows = stats["by_surgeon"]
    if not rows:
        _show_no_data(); return None
    names = [r[0].replace("Dr. ", "") for r in rows]
    counts = [r[1] for r in rows]
    fig, ax = _base_fig("Surgeries per Surgeon", figsize=(10, 5))
    bars = ax.barh(names, counts, color=PALETTE[:len(names)], edgecolor="#0f1923", height=0.6)
    for bar, val in zip(bars, counts):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2,
                str(val), va="center", color="#e8f4f8", fontsize=10)
    ax.set_xlabel("Number of Surgeries")
    ax.invert_yaxis()
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, color="#e8f4f8")
    plt.tight_layout()
    return fig


def chart_surgeries_by_month(parent_frame=None):
    stats = db.get_surgeries_stats()
    rows = stats["by_month"]
    if not rows:
        _show_no_data(); return None
    months = [r[0] for r in rows]
    counts = [r[1] for r in rows]
    fig, ax = _base_fig("Monthly Surgery Volume", figsize=(10, 5))
    ax.fill_between(months, counts, alpha=0.3, color="#2eafd4")
    ax.plot(months, counts, marker="o", color="#2eafd4", linewidth=2.5,
            markersize=8, markerfacecolor="#e8a838")
    for x, y in zip(months, counts):
        ax.annotate(str(y), (x, y), textcoords="offset points",
                    xytext=(0, 8), ha="center", color="#e8f4f8", fontsize=9)
    ax.set_xlabel("Month")
    ax.set_ylabel("Surgeries")
    plt.xticks(rotation=30, ha="right", color="#a0b8c8")
    plt.tight_layout()
    return fig


def chart_priority_distribution(parent_frame=None):
    stats = db.get_surgeries_stats()
    data = stats["by_priority"]
    if not data:
        _show_no_data(); return None
    order = ["Emergency", "High", "Normal", "Low"]
    labels = [k for k in order if k in data]
    values = [data[k] for k in labels]
    colors_map = {"Emergency":"#e05252","High":"#e8a838","Normal":"#4caf82","Low":"#2eafd4"}
    colors = [colors_map[l] for l in labels]
    fig, ax = _base_fig("Surgery Priority Distribution")
    bars = ax.bar(labels, values, color=colors, edgecolor="#0f1923", width=0.5)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                str(val), ha="center", color="#e8f4f8", fontsize=11, fontweight="bold")
    ax.set_ylabel("Count")
    ax.set_xticklabels(labels, color="#e8f4f8")
    plt.tight_layout()
    return fig


def chart_surgery_types(parent_frame=None):
    stats = db.get_surgeries_stats()
    rows = stats["by_type"]
    if not rows:
        _show_no_data(); return None
    types = [r[0] for r in rows]
    counts = [r[1] for r in rows]
    fig, ax = _base_fig("Top Surgery Types", figsize=(10, 5))
    palette = sns.color_palette(PALETTE, len(types))
    bars = ax.barh(types, counts, color=palette, edgecolor="#0f1923", height=0.6)
    for bar, val in zip(bars, counts):
        ax.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height()/2,
                str(val), va="center", color="#e8f4f8", fontsize=10)
    ax.invert_yaxis()
    ax.set_xlabel("Count")
    ax.set_yticklabels(types, color="#e8f4f8")
    plt.tight_layout()
    return fig


def _show_no_data():
    fig, ax = _base_fig("No Data Available")
    ax.text(0.5, 0.5, "No data to display.\nAdd surgeries first.",
            ha="center", va="center", color="#a0b8c8", fontsize=14,
            transform=ax.transAxes)
    plt.tight_layout()
    return fig


def _stamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")
